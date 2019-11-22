#!/usr/bin/env python3

"""Insert the contents of multiple text files into a single spreadsheet."""

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


path = './chap_12/text2sheet'
file_list = ['test.txt', 'test2.txt']

wb = openpyxl.Workbook()
sheet = wb.active

column_num = 1
for file in file_list:
    # Open file and format lines
    lines = open(path + '/' + file).readlines()

    make_bold = Font(bold=True)
    sheet.cell(row=1, column=column_num).value = file
    sheet.cell(row=1, column=column_num).font = make_bold

    longest = 0
    row_num = 2
    for line in lines:
        line = line.strip()

        # Calculate appropriate column width
        if len(line) > longest:
            longest = len(line)

        # Write lines to spreadsheet
        sheet.cell(row=row_num, column=column_num).value = line
        row_num += 1

    column_letter = get_column_letter(column_num)
    sheet.column_dimensions[column_letter].width = longest
    column_num += 1

wb.save(path + '/text2sheet.xlsx')

print("Spreadsheet saved as text2sheet.xlsx - it can be found in the same "
      "directory as the inputted files.")
